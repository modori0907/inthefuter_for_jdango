from datetime import datetime, timedelta
# コメントを記述した行を最新にしたい場合
from django.db.models import Subquery, OuterRef
from django.utils import timezone


# excelダウンロード用
import openpyxl
# Create your views here.
# csv ダウンロード用
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
# 詳細画面を表示するため
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView

# 登録画面を作成するため
from .forms import PatchForm, CommentForm
from .models import (
    Patchs, Comment
)

# 検索画面を表示
def index(request):
    return render(request, 'patch/patch_index.html')

# パッチリストを作成する為の処理
def patch_create(request):
    if request.method == 'POST':
        form = PatchForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('patch_list:list')

    else:
        form = PatchForm()
    return render(request, 'patch/patch_create.html', {'form': form})


# パッチリストを更新する為の処理
def patch_update(request, pk):
    patch = Patchs.objects.get(pk=pk)
    if request.method == 'POST':
        form = PatchForm(request.POST, instance=patch)
        if form.is_valid():
            form.save()
            return redirect('patch_list:list')
    else:
        form = PatchForm(instance=patch)
    return render(request, 'patch/patch_update.html', {'form': form})




# パッチリストを削除する処理
def patch_delete(request, pk):
    patch = Patchs.objects.get(pk=pk)
    patch.delete()
    return redirect('patch_list:list')


# パッチリストの詳細画面を表示する
# コメント登録機能を追加

class PatchDetailView(DetailView):
    model = Patchs
    template_name = 'patch/patch_detail.html'
    context_object_name = 'patch_list'

    # 閲覧した回数をカウントする為の関数
    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        #　閲覧回数のインクリメント
        obj.views_count += 1
        obj.save()
        return super().get(request, *args, **kwargs)



    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['comments'] = Comment.objects.filter(patchs=self.get_object())
        context['form'] = CommentForm()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = CommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.patchs = self.object
            comment.save()
            return redirect('patch_list:detail', pk=self.object.pk)
        context = self.get_context_data(object=self.object, form=form)
        return self.render_to_response(context)


class PatchListView(ListView):
    # modelで作成したclassを指定
    model = Patchs
    template_name = 'patch/patch_list.html'

    # 検索画面の結果を表示させるため。getで取得した値
    def get_queryset(self):
        query = super().get_queryset()
        # URLに記載した名前
        name = self.request.GET.get('application_name', None)
        # checks = self.request.GET.get('patch_check', None)
        impact_check = self.request.GET.get('impact_check', None)

        start_month = self.request.GET.get('start_date', None)
        end_month = self.request.GET.get('end_date', None)

        if name:
            query = query.filter(
                name=name
            )

        if impact_check:
            query = query.filter(
                impact_check=impact_check
            )

        if start_month and end_month:
            start_month = datetime.strptime(start_month, '%Y-%m')
            end_month = datetime.strptime(end_month, '%Y-%m')
            query = query.filter(release_date__range=(start_month, end_month))

        # サブクエリを利用して、最新のコメントの日時を取得する
        subquery = Comment.objects.filter(patchs_id=OuterRef('pk')).order_by('-created_date')
        query = query.annotate(latest_comment=Subquery(subquery.values('created_date')[:1]))

        # 最新のコメントがあるパッチを上位に表示する
        query = query.order_by('-latest_comment')

        # 更新および新規作成を判定する
        today = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)

        for obj in query:

            # 1日より前に作成されたもの
            if obj.created_at >= today - timedelta(days=1):
                obj.is_new = True
            else:
                obj.is_new = False
            if obj.updated_at >= today - timedelta(days=1):
                obj.is_updated = True
            else:
                obj.is_updated = False

        return query

    # csvダウンロード用追加した
    # テンプレートに渡すコンテキストデータを返すメソッド
    # ListView クラスのget_context_data メソッドを利用
    def get_context_data(self, **kwargs):
        # クラスのget_context_dataメソッドを呼び出し、コンテキストデータを取得
        context = super().get_context_data(**kwargs)

        # GETパラメーターにapplication_nameが含まれている場合に、CSVファイルのダウンロードURLをコンテキストに追加
        # ここで指定するのはhtmlで記載された番号をしていする
        if 'application_name' or 'impact_check' or 'start_date' or 'end_date' in self.request.GET:
            # reverse('patch_list:list')で、patch_listという名前のURLパターンのURLを取得
            # self.request.GET.urlencode()で、GETパラメーターをエンコードした文字列を取得
            context['excel_url'] = reverse('patch_list:list') + '?' + self.request.GET.urlencode()
            # context['csv_url']に、CSVファイルのダウンロードURLを追加
            context['excel_url'] += '&export=excel'
        return context

    # ここは表示している内容によって変更できるようにする必要がある
    # エクセス記載処理
    def create_excel_response(self, queryset):
        response = HttpResponse(content_type='application/vnd.ms-excel')
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        response['Content-Disposition'] = f'attachment; filename="search_results_{timestamp}.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        worksheet['A1'] = 'Name'
        worksheet['B1'] = 'Patch Name'
        worksheet['C1'] = 'Patch No'
        worksheet['D1'] = 'Release Date'
        worksheet['E1'] = 'Content'

        row_num = 2
        for patch in queryset:
            worksheet.cell(row=row_num, column=1, value=patch.name)
            worksheet.cell(row=row_num, column=2, value=patch.patch_name)
            worksheet.cell(row=row_num, column=3, value=patch.patch_no)
            worksheet.cell(row=row_num, column=4, value=patch.release_date)
            row_num += 1

        workbook.save(response)

        return response

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET and request.GET['export'] == 'excel':
            queryset = self.get_queryset()
            response = self.create_excel_response(queryset)
            return response
        else:
            return super().get(request, *args, **kwargs)
